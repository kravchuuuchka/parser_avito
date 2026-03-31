"""
exporter_config.py - Настройки внешнего вида результирующего Excel-файла
"""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HEADER_FONT = Font(name="Arial", bold=True, color="000000", size=11)
CELL_FONT = Font(name="Arial", size=10)

HEADER_FILL = PatternFill(fill_type=None)
STATUS_ACTIVE_FILL = PatternFill(fill_type="solid", fgColor="E2EFDA")
STATUS_CLOSED_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")

HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=False)

_SIDE = Side(style="thin", color="000000")
BORDER = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)

COLUMNS = [
    ("ID", "id", 18),
    ("Название", "title", 40),
    ("Цена (руб.)", "price", 15),
    ("Адрес", "address", 30),
    ("Описание", "description", 50),
    ("Дата публикации", "published_on", 18),
    ("Просмотры", "views", 12),
    ("Статус", "status", 12),
    ("Город", "city", 18),
    ("Запрос", "query", 18),
    ("Ссылка", "url", 50),
    ("Добавлен в кэш", "cached_at", 20),
    ("Обновлён в кэше", "updated_at", 20),
]

HEADER_ROW_HEIGHT = 30
