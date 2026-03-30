"""
exporter.py - Экспорт объявлений из кэша в Excel (.xlsx)
"""

from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from loguru import logger

from core.ad import Ad
from exporter.exporter_config import *

def _format_status(status: int) -> str:
    return "Активно" if status == 1 else "Закрыто"


def _get_value(ad: Ad, key: str):
    value = getattr(ad, key)
    if key == "status":
        return _format_status(value)
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    return value if value is not None else ""


def export(ads: list[Ad], filepath: str) -> None:
    """
    Экспорт списка объявлений в Excel-файл

    Args:
        ads:      Список объектов Ad
        filepath: Путь к выходному .xlsx файлу
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Объявления"

    ws.append([col[0] for col in COLUMNS])
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border    = BORDER
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

    status_col_idx = [col[1] for col in COLUMNS].index("status") + 1

    for ad in ads:
        ws.append([_get_value(ad, key) for _, key, _ in COLUMNS])
        row_idx = ws.max_row
        for col_idx, cell in enumerate(ws[row_idx], 1):
            cell.font      = CELL_FONT
            cell.alignment = CELL_ALIGNMENT
            cell.border    = BORDER
            if col_idx == status_col_idx:
                cell.fill = STATUS_ACTIVE_FILL if ad.status == 1 else STATUS_CLOSED_FILL

    for col_idx, (_, _, width) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    wb.save(filepath)
    logger.info("Экспортировано {} объявлений в {}", len(ads), filepath)