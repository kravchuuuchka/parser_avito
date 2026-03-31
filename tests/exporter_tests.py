"""
exporter_tests.py - Тесты для exporter.py
Читает БД, созданную в cache_tests (C:\test\\cache_test.db)
Excel сохраняется в C:\\test\\exporter_test.xlsx
"""

import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, os.path.dirname(__file__))

import sqlite3

from helpers import check, ensure_test_dir, summary
from loguru import logger
from openpyxl import load_workbook

import exporter.exporter as exporter
from core.cache import _row_to_ad, _str_to_date, _str_to_dt


def run():
    logger.info("====exporter====")

    test_dir = ensure_test_dir()
    db = os.path.join(test_dir, "cache_test.db")
    xlsx = os.path.join(test_dir, "exporter_test.xlsx")

    if not os.path.exists(db):
        logger.error("БД не найдена: {}. Сначала запусти cache_tests.", db)
        return

    with sqlite3.connect(db) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("SELECT * FROM ads").fetchall()

    ads = [_row_to_ad(row) for row in rows]
    exporter.export(ads, xlsx)

    wb = load_workbook(xlsx)
    ws = wb.active

    from exporter.exporter_config import COLUMNS

    check("файл создан", os.path.exists(xlsx), True)
    check("количество строк", ws.max_row, len(ads) + 1)

    for col_idx, (header, _, _w) in enumerate(COLUMNS, 1):
        check(
            f"заголовок колонки {col_idx}: «{header}»",
            ws.cell(1, col_idx).value,
            header,
        )

    excel_rows = {ws.cell(r, 1).value: r for r in range(2, ws.max_row + 1)}

    for row in rows:
        ad_id = row["id"]
        if ad_id not in excel_rows:
            logger.error("Объявление {} не найдено в Excel", ad_id)
            continue

        r = excel_rows[ad_id]

        expected_price = row["price"]
        check(f"{ad_id}: цена {expected_price}", ws.cell(r, 3).value, expected_price)

        expected_status = "Активно" if row["status"] == 1 else "Закрыто"
        check(
            f"{ad_id}: статус - {expected_status}", ws.cell(r, 8).value, expected_status
        )

        published_on = _str_to_date(row["published_on"])
        if published_on:
            expected_date = published_on.strftime("%d.%m.%Y")
            check(
                f"{ad_id}: дата публикации - {expected_date}",
                ws.cell(r, 6).value,
                expected_date,
            )

        cached_at_val = ws.cell(r, 12).value
        updated_at_val = ws.cell(r, 13).value
        check(f"{ad_id}: cached_at заполнен", cached_at_val is not None, True)
        check(f"{ad_id}: updated_at заполнен", updated_at_val is not None, True)

        cached_dt = _str_to_dt(row["cached_at"])
        if cached_dt:
            expected_cached = cached_dt.strftime("%d.%m.%Y %H:%M:%S")
            check(
                f"{ad_id}: cached_at отформатирован верно",
                cached_at_val,
                expected_cached,
            )

    wb.close()
    logger.info("Excel сохранён: {}", xlsx)


if __name__ == "__main__":
    run()
    summary()
